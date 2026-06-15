#!/usr/bin/env python3
"""Read and analyse Excel (.xlsx / .xls / .csv) files.

Usage:
    python read.py workbook.xlsx
    python read.py workbook.xlsx --sheet "Q2"
    python read.py workbook.xlsx --sheets          # list all sheet names
    python read.py workbook.xlsx --stats           # summary statistics
    python read.py workbook.xlsx --cell B5         # read a single cell
    python read.py workbook.xlsx --metadata
"""

import json
import sys


def list_sheets(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  {name}  ({ws.max_row} rows × {ws.max_column} cols)")
    wb.close()


def read_sheet(path, sheet=None):
    import pandas as pd

    df = pd.read_excel(path, sheet_name=sheet or 0)
    print(df.to_string(index=False))


def show_stats(path, sheet=None):
    import pandas as pd

    df = pd.read_excel(path, sheet_name=sheet or 0)
    print(df.describe(include="all").to_string())


def read_cell(path, cell_ref, sheet=None):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    print(ws[cell_ref].value)
    wb.close()


def show_metadata(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    props = wb.properties
    print(
        json.dumps(
            {
                "title": props.title or "",
                "subject": props.subject or "",
                "creator": props.creator or "",
                "keywords": props.keywords or "",
                "created": str(props.created or ""),
                "modified": str(props.modified or ""),
                "sheets": wb.sheetnames,
            },
            indent=2,
        )
    )
    wb.close()


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] in {"-h", "--help"}:
        print(__doc__)
        sys.exit(0)

    path = args[0]
    sheet = None
    if "--sheet" in args:
        sheet = args[args.index("--sheet") + 1]

    if "--sheets" in args:
        list_sheets(path)
    elif "--stats" in args:
        show_stats(path, sheet)
    elif "--cell" in args:
        cell_ref = args[args.index("--cell") + 1]
        read_cell(path, cell_ref, sheet)
    elif "--metadata" in args:
        show_metadata(path)
    else:
        read_sheet(path, sheet)
