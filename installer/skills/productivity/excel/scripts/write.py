#!/usr/bin/env python3
"""Create and edit Excel (.xlsx) workbooks.

Usage:
    # Create a workbook from a CSV file
    python write.py --from-csv data.csv output.xlsx

    # Update a single cell
    python write.py --set-cell B5 "=SUM(B2:B4)" workbook.xlsx
    python write.py --set-cell B5 "=SUM(B2:B4)" workbook.xlsx --sheet "Q2"

    # Append a row of values
    python write.py --append-row "April,120000,85000" workbook.xlsx

    # Export a sheet to CSV
    python write.py --to-csv workbook.xlsx output.csv
    python write.py --to-csv workbook.xlsx output.csv --sheet "Q2"

    # Convert to PDF via LibreOffice
    python write.py --to-pdf workbook.xlsx
"""

import subprocess
import sys
from pathlib import Path


def from_csv(src, dst):
    import pandas as pd

    df = pd.read_csv(src)
    df.to_excel(dst, index=False)
    print(f"Created {dst} from {src} ({len(df)} rows)")


def set_cell(path, cell_ref, value, sheet=None):
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    ws[cell_ref] = value
    wb.save(path)
    print(f"Set {cell_ref} = {value!r} in {path}")


def append_row(path, row_csv, sheet=None):
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    # parse simple CSV values; try int/float coercion
    values = []
    for v in row_csv.split(","):
        v = v.strip()
        try:
            values.append(int(v))
        except ValueError:
            try:
                values.append(float(v))
            except ValueError:
                values.append(v)
    ws.append(values)
    wb.save(path)
    print(f"Appended row to {path}: {values}")


def to_csv(path, out, sheet=None):
    import pandas as pd

    df = pd.read_excel(path, sheet_name=sheet or 0)
    df.to_csv(out, index=False, encoding="utf-8-sig")
    print(f"Exported {out} ({len(df)} rows)")


def to_pdf(path):
    path = str(Path(path).resolve())
    out = Path(path).with_suffix(".pdf")
    result = subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to",
            "pdf",
            str(path),
            "--outdir",
            str(Path(path).parent),
        ],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        print(f"Error: {result.stderr}", file=sys.stderr)
        sys.exit(1)
    print(f"PDF: {out}")


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] in {"-h", "--help"}:
        print(__doc__)
        sys.exit(0)

    sheet = None
    if "--sheet" in args:
        sheet = args[args.index("--sheet") + 1]

    if "--from-csv" in args:
        idx = args.index("--from-csv")
        from_csv(args[idx + 1], args[idx + 2])

    elif "--set-cell" in args:
        idx = args.index("--set-cell")
        cell_ref = args[idx + 1]
        value = args[idx + 2]
        path = args[idx + 3]
        set_cell(path, cell_ref, value, sheet)

    elif "--append-row" in args:
        idx = args.index("--append-row")
        row_csv = args[idx + 1]
        path = args[idx + 2]
        append_row(path, row_csv, sheet)

    elif "--to-csv" in args:
        idx = args.index("--to-csv")
        path = str(Path(args[idx + 1]).resolve())
        out = (
            args[idx + 2]
            if idx + 2 < len(args) and not args[idx + 2].startswith("--")
            else Path(path).with_suffix(".csv")
        )
        to_csv(path, str(out), sheet)

    elif "--to-pdf" in args:
        idx = args.index("--to-pdf")
        path = args[idx + 1]
        to_pdf(path)

    else:
        print("Unknown arguments. Run with --help for usage.")
        sys.exit(1)
